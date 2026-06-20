"""
Excel fill engine for Customs Operations — 截关表格生成
Five fillers: ENS, ICS2, MultiProduct, Manifest, LoadingNotice
Uses openpyxl to fill templates while preserving formatting.
"""

from __future__ import annotations

import copy
import shutil
import zipfile
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter

# ── Resolve paths ──
BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = BASE_DIR / "templates"
OUTPUT_DIR = BASE_DIR / "data" / "output"


def _copy_template(name: str, job_id: int) -> tuple[Path, openpyxl.Workbook]:
    """Copy a template to the output dir and return (path, workbook)."""
    src = TEMPLATE_DIR / name
    if not src.exists():
        raise FileNotFoundError(f"Template not found: {src}")
    out_dir = OUTPUT_DIR / str(job_id)
    out_dir.mkdir(parents=True, exist_ok=True)
    dst = out_dir / name
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)
    return dst, wb


def _set_cell(ws, coord: str, value):
    """Set cell value. Handles merged cells by writing to top-left."""
    # If value is empty/None, skip (don't clear template labels)
    if value is None or value == "":
        return
    cell = ws[coord]
    # Check if this cell is part of a merged range
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # Write to the top-left cell of the merged range
            tl = merged_range.min_row, merged_range.min_col
            ws.cell(row=tl[0], column=tl[1], value=value)
            return
    cell.value = value


class ENSFiller:
    """Fill 空模板_ENS_VGM申报表.xlsx — ENS / VGM Declaration."""

    TEMPLATE = "空模板_ENS_VGM申报表.xlsx"

    def __init__(self, job):
        self.job = job

    def fill(self) -> Path:
        path, wb = _copy_template(self.TEMPLATE, self.job.id)
        ws = wb.active

        # Contact info
        _set_cell(ws, "E2", self.job.ens_contact_person)
        _set_cell(ws, "E3", self.job.ens_contact_email)
        _set_cell(ws, "E4", self.job.ens_contact_phone)
        _set_cell(ws, "E5", self.job.ens_contact_fax)

        # B/L numbers
        _set_cell(ws, "B6", self.job.mbl_no)
        _set_cell(ws, "E6", self.job.booking_no)

        # Shipper (merged B7:C11)
        shipper_text = "\n".join(filter(None, [
            self.job.shipper_name,
            self.job.shipper_address,
            f"TEL: {self.job.shipper_phone}" if self.job.shipper_phone else "",
            f"FAX: {self.job.shipper_fax}" if self.job.shipper_fax else "",
        ]))
        _set_cell(ws, "B7", shipper_text)

        # Consignee (merged B12:C16)
        consignee_text = "\n".join(filter(None, [
            self.job.consignee_name,
            self.job.consignee_address,
            f"TEL: {self.job.consignee_phone}" if self.job.consignee_phone else "",
            f"FAX: {self.job.consignee_fax}" if self.job.consignee_fax else "",
        ]))
        _set_cell(ws, "B12", consignee_text)

        # Notify Party (merged B17:C21)
        notify_text = "\n".join(filter(None, [
            self.job.notifier_name,
            self.job.notifier_address,
            f"TEL: {self.job.notifier_phone}" if self.job.notifier_phone else "",
            f"FAX: {self.job.notifier_fax}" if self.job.notifier_fax else "",
        ]))
        _set_cell(ws, "B17", notify_text)

        # Vessel / Voyage
        _set_cell(ws, "B22", f"{self.job.vessel_name} / {self.job.voyage}")
        _set_cell(ws, "B23", self.job.pol)
        _set_cell(ws, "E23", self.job.place_of_receipt)
        _set_cell(ws, "B24", self.job.pod)
        _set_cell(ws, "E24", self.job.place_of_delivery)

        # Container type/qty
        container_info = self._build_container_summary()
        _set_cell(ws, "D25", container_info)

        # Goods description
        _set_cell(ws, "C26", self.job.ens_goods_desc)

        wb.save(path)
        wb.close()
        return path

    def _build_container_summary(self) -> str:
        """Summarise container types and quantities."""
        if not self.job.containers:
            return ""
        type_counts: dict[str, int] = {}
        for c in self.job.containers:
            ct = c.get("container_type", "") or "Unknown"
            type_counts[ct] = type_counts.get(ct, 0) + 1
        parts = [f"{v}×{k}" for k, v in type_counts.items()]
        return " / ".join(parts)


class ICS2Filler:
    """Fill 空模板_ICS2舱单数据表.xlsx — ICS2 EU Customs Manifest.

    Supports both F15 (with HBL) and F17 (without HBL) declaration types.
    """

    TEMPLATE = "空模板_ICS2舱单数据表.xlsx"

    def __init__(self, job):
        self.job = job

    def fill(self) -> Path:
        path, wb = _copy_template(self.TEMPLATE, self.job.id)
        ws = wb.active

        is_f15 = self.job.ics2_declaration_type == "F15"

        # Row 3 — MBL info
        _set_cell(ws, "B3", self.job.mbl_no)
        _set_cell(ws, "C3", self.job.mbl_contract_no)
        _set_cell(ws, "F3", self.job.mbl_type)
        _set_cell(ws, "H3", self.job.payment_type)
        _set_cell(ws, "J3", self.job.container_mark)

        # Row 4 — Declaration type
        decl_label = "F15-有HBL" if self.job.has_hbl else "F17-无HBL"
        _set_cell(ws, "B4", decl_label)
        _set_cell(ws, "C4", self.job.hbl_contract_no)
        if is_f15:
            _set_cell(ws, "F4", self.job.hbl_type)
        _set_cell(ws, "H4", self.job.transport_mode)

        # Row 5
        if is_f15:
            _set_cell(ws, "B5", self.job.hbl_no)
        _set_cell(ws, "D5", self.job.carrier)
        _set_cell(ws, "F5", self.job.ics2_member_state)
        if is_f15:
            _set_cell(ws, "H5", self.job.pol)
            _set_cell(ws, "J5", self.job.pod)

        # Row 6 — Weights
        _set_cell(ws, "B6", self.job.mbl_total_weight)
        if is_f15:
            _set_cell(ws, "D6", self.job.hbl_total_weight)
        _set_cell(ws, "E6", self.job.imo)
        _set_cell(ws, "H6", self.job.voyage)

        # Row 8 — Transit countries
        _set_cell(ws, "B8", self.job.transit_countries)

        # Rows 9-18 — Shipper / Consignee
        if is_f15:
            self._fill_party(ws, "B", self.job, "shipper")
            self._fill_party(ws, "F", self.job, "consignee")

        # Rows 19-28 — Seller / Buyer
        self._fill_party(ws, "B", self.job, "seller")
        self._fill_party(ws, "F", self.job, "buyer")

        # Rows 29-37 — ICS2 Declarant
        _set_cell(ws, "C29", self.job.ics2_declarant_eori)
        _set_cell(ws, "C30", self.job.ics2_declarant_name)
        _set_cell(ws, "C31", self.job.ics2_declarant_country_code)
        _set_cell(ws, "C32", self.job.ics2_declarant_city)
        _set_cell(ws, "C33", self.job.ics2_declarant_street)
        _set_cell(ws, "C34", self.job.ics2_declarant_street_no)
        _set_cell(ws, "C35", self.job.ics2_declarant_postal_code)
        _set_cell(ws, "C36", self.job.ics2_declarant_po_box)
        _set_cell(ws, "C37", self.job.ics2_declarant_phone)

        # Row 38+ — Container / Product / Package detail (F15 only)
        if is_f15:
            self._fill_containers(ws, self.job.containers, row=40)
            self._fill_products(ws, self.job.products, row=44)
            self._fill_packages(ws, self.job.products, row=self._pkg_start_row(self.job))

        wb.save(path)
        wb.close()
        return path

    def _fill_party(self, ws, col_prefix: str, job, role: str):
        """Fill a party block (Shipper/Consignee/Seller/Buyer)."""
        prefix = {"shipper": "shipper", "consignee": "consignee",
                   "seller": "seller", "buyer": "buyer"}[role]
        _, eori, name, ptype, country, city, street, street_no, postal, po_box, phone = (
            getattr(job, f"{prefix}_eori", ""),
            getattr(job, f"{prefix}_eori", ""),
            getattr(job, f"{prefix}_name", ""),
            getattr(job, f"{prefix}_type", ""),
            getattr(job, f"{prefix}_country_code", ""),
            getattr(job, f"{prefix}_city", ""),
            getattr(job, f"{prefix}_street", ""),
            getattr(job, f"{prefix}_street_no", ""),
            getattr(job, f"{prefix}_postal_code", ""),
            getattr(job, f"{prefix}_po_box", ""),
            getattr(job, f"{prefix}_phone", ""),
        )

        if role in ("shipper", "seller"):
            c = col_prefix  # B column
        else:
            c = "G" if col_prefix == "F" else col_prefix

        # EORI row
        if role in ("shipper", "consignee"):
            base = 9
            _set_cell(ws, f"{col_prefix}9", eori)
            _set_cell(ws, f"{c}10", name)
            _set_cell(ws, f"{c}11", ptype)
            _set_cell(ws, f"{c}12", country)
            _set_cell(ws, f"{c}13", city)
            _set_cell(ws, f"{c}14", street)
            _set_cell(ws, f"{c}15", street_no)
            _set_cell(ws, f"{c}16", postal)
            _set_cell(ws, f"{c}17", po_box)
            _set_cell(ws, f"{c}18", phone)
        else:  # seller / buyer
            base = 19
            _set_cell(ws, f"{col_prefix}19", eori)
            _set_cell(ws, f"{c}20", name)
            _set_cell(ws, f"{c}21", ptype)
            _set_cell(ws, f"{c}22", country)
            _set_cell(ws, f"{c}23", city)
            _set_cell(ws, f"{c}24", street)
            _set_cell(ws, f"{c}25", street_no)
            _set_cell(ws, f"{c}26", postal)
            _set_cell(ws, f"{c}27", po_box)
            _set_cell(ws, f"{c}28", phone)

    def _fill_containers(self, ws, containers: list, row: int):
        """Fill container detail rows dynamically."""
        if not containers:
            return
        start_row = row
        for i, c in enumerate(containers):
            r = start_row + i
            if r > start_row:
                # Insert row preserving format
                ws.insert_rows(r)
                self._copy_row_style(ws, start_row, r)
            _set_cell(ws, f"B{r}", c.get("container_no", ""))
            _set_cell(ws, f"C{r}", c.get("seal_no", ""))
            _set_cell(ws, f"D{r}", c.get("container_type", ""))
            _set_cell(ws, f"E{r}", c.get("is_soc", ""))
            _set_cell(ws, f"F{r}", c.get("status", ""))

    def _fill_products(self, ws, products: list, row: int):
        """Fill product info rows."""
        if not products:
            return
        start_row = row
        for i, p in enumerate(products):
            r = start_row + i
            if r > start_row:
                ws.insert_rows(r)
                self._copy_row_style(ws, start_row, r)
            _set_cell(ws, f"B{r}", i + 1)  # seq
            _set_cell(ws, f"C{r}", p.get("description", ""))
            _set_cell(ws, f"D{r}", p.get("hs_code", ""))
            # Container nos for this product (| separated)
            container_nos = p.get("container_nos", "")
            _set_cell(ws, f"E{r}", container_nos)
            _set_cell(ws, f"F{r}", p.get("weight", 0))
            _set_cell(ws, f"G{r}", p.get("cus_code", ""))
            _set_cell(ws, f"H{r}", p.get("undg", ""))

    def _pkg_start_row(self, job) -> int:
        """Determine the row where package info starts (after products)."""
        base = 47
        if job.containers:
            base += max(len(job.containers) - 1, 0)
        if job.products:
            base += max(len(job.products) - 1, 0)
        return base

    def _fill_packages(self, ws, products: list, row: int):
        """Fill package count rows."""
        if not products:
            return
        for i, p in enumerate(products):
            r = row + i
            if r > row:
                ws.insert_rows(r)
                self._copy_row_style(ws, row, r)
            _set_cell(ws, f"B{r}", i + 1)  # seq matching product
            _set_cell(ws, f"C{r}", p.get("packages", 0))
            _set_cell(ws, f"D{r}", p.get("pkg_unit", ""))
            _set_cell(ws, f"E{r}", p.get("marks", "N/M"))

    @staticmethod
    def _copy_row_style(ws, src_row: int, dst_row: int):
        """Copy cell formatting from source row to destination row."""
        for col in range(1, ws.max_column + 1):
            src = ws.cell(row=src_row, column=col)
            dst = ws.cell(row=dst_row, column=col)
            if src.has_style:
                dst.font = copy.copy(src.font)
                dst.border = copy.copy(src.border)
                dst.fill = copy.copy(src.fill)
                dst.number_format = src.number_format
                dst.protection = copy.copy(src.protection)
                dst.alignment = copy.copy(src.alignment)


class MultiProductFiller:
    """Fill 空模板_多品名表.xlsx — Multi-product table (2+ products only)."""

    TEMPLATE = "空模板_多品名表.xlsx"

    def __init__(self, job):
        self.job = job

    def fill(self) -> Path:
        path, wb = _copy_template(self.TEMPLATE, self.job.id)
        ws = wb.active

        # Bill No
        _set_cell(ws, "A3", self.job.mbl_no or self.job.booking_no)

        # Product rows (starts at row 4, up to row 8)
        products = self.job.products or []
        for i, p in enumerate(products[:5]):
            r = 4 + i
            _set_cell(ws, f"D{r}", p.get("packages", ""))
            _set_cell(ws, f"E{r}", p.get("pkg_unit", ""))
            _set_cell(ws, f"F{r}", p.get("weight", ""))
            _set_cell(ws, f"G{r}", "")  # M3 — keep empty unless known
            _set_cell(ws, f"H{r}", p.get("description", ""))
            _set_cell(ws, f"I{r}", p.get("hs_code", ""))
            dg = "DG" if p.get("undg") else "NON-DG"
            _set_cell(ws, f"J{r}", dg)

        wb.save(path)
        wb.close()
        return path


class ManifestFiller:
    """Fill 空模板_DEHAM舱单表.xlsx — Cargo Manifest (DE-HAM format)."""

    TEMPLATE = "空模板_DEHAM舱单表.xlsx"

    def __init__(self, job):
        self.job = job

    def fill(self) -> Path:
        path, wb = _copy_template(self.TEMPLATE, self.job.id)
        ws = wb.active

        # Shipper
        self._fill_party(ws, "Shipper", "shipper", row_start=2)
        # Consignee
        self._fill_party(ws, "Consignee", "consignee", row_start=11)
        # Notifier
        self._fill_party(ws, "Notifier", "notifier", row_start=22)

        # Vessel / Voyage
        _set_cell(ws, "B32", self.job.vessel_name)
        _set_cell(ws, "B33", self.job.voyage)
        _set_cell(ws, "H32", self.job.pol)
        _set_cell(ws, "H33", self.job.pod)

        # Cargo detail rows (starting at row 37)
        products = self.job.products or []
        containers = self.job.containers or []
        # Combine container and product data into detail rows
        rows_data = self._build_detail_rows(containers, products)
        start_row = 37
        for i, row_data in enumerate(rows_data):
            r = start_row + i
            if r > start_row:
                ws.insert_rows(r)
                self._copy_row_style(ws, start_row, r)
            for col_idx, val in enumerate(row_data, start=1):
                _set_cell(ws, f"{get_column_letter(col_idx)}{r}", val)

        wb.save(path)
        wb.close()
        return path

    def _fill_party(self, ws, label: str, role: str, row_start: int):
        """Fill a party block (Shipper/Consignee/Notifier)."""
        prefix = role
        r = row_start
        _set_cell(ws, f"D{r}", getattr(self.job, f"{prefix}_code", ""))
        _set_cell(ws, f"D{r+1}", getattr(self.job, f"{prefix}_name", ""))
        _set_cell(ws, f"D{r+2}", getattr(self.job, f"{prefix}_address", ""))
        _set_cell(ws, f"D{r+3}", getattr(self.job, f"{prefix}_country_code", ""))
        _set_cell(ws, f"D{r+4}", getattr(self.job, f"{prefix}_phone", ""))
        _set_cell(ws, f"D{r+5}", getattr(self.job, f"{prefix}_fax", ""))
        _set_cell(ws, f"D{r+6}", getattr(self.job, f"{prefix}_email", ""))
        _set_cell(ws, f"D{r+7}", getattr(self.job, f"{prefix}_aeo", ""))
        if role == "consignee":
            _set_cell(ws, f"D{r+8}", getattr(self.job, "consignee_contact_person", ""))
            _set_cell(ws, f"D{r+9}", getattr(self.job, "consignee_contact_phone", ""))

    def _build_detail_rows(self, containers: list, products: list) -> list[list]:
        """Build combined container+product detail rows for the manifest."""
        rows = []
        if containers and products:
            # Match by index — each container row has corresponding product info
            for idx, c in enumerate(containers):
                p = products[idx] if idx < len(products) else {}
                row = [
                    c.get("container_no", ""),       # A: 箱号
                    c.get("seal_no", ""),             # B: 封号
                    c.get("container_type", ""),      # C: 箱型
                    self.job.mbl_no,                  # D: 提单号
                    p.get("packages", ""),            # E: 件数
                    p.get("pkg_unit", ""),            # F: 包装单位
                    p.get("weight", ""),              # G: 毛重(KGS)
                    "",                               # H: 体积(CBM)
                    p.get("description", ""),         # I: 英文品名
                    p.get("hs_code", ""),             # J: 10位Hscode
                ]
                rows.append(row)
        elif products:
            for p in products:
                rows.append([
                    "", "", "",
                    self.job.mbl_no,
                    p.get("packages", ""),
                    p.get("pkg_unit", ""),
                    p.get("weight", ""),
                    "",
                    p.get("description", ""),
                    p.get("hs_code", ""),
                ])
        return rows or [["", "", "", "", "", "", "", "", "", ""]]

    @staticmethod
    def _copy_row_style(ws, src_row: int, dst_row: int):
        for col in range(1, ws.max_column + 1):
            src = ws.cell(row=src_row, column=col)
            dst = ws.cell(row=dst_row, column=col)
            if src.has_style:
                dst.font = copy.copy(src.font)
                dst.border = copy.copy(src.border)
                dst.fill = copy.copy(src.fill)
                dst.number_format = src.number_format
                dst.protection = copy.copy(src.protection)
                dst.alignment = copy.copy(src.alignment)


class LoadingNoticeFiller:
    """Fill 空模板_做箱通知.xlsx — Factory Loading Notice."""

    TEMPLATE = "空模板_做箱通知.xlsx"

    def __init__(self, job):
        self.job = job

    def fill(self) -> Path:
        path, wb = _copy_template(self.TEMPLATE, self.job.id)
        ws = wb.active

        _set_cell(ws, "C8", self.job.receiving_company)
        fm = f"JAS / {self.job.fm_department}" if self.job.fm_department else "JAS / "
        _set_cell(ws, "C9", fm)
        _set_cell(ws, "C10", self.job.cc_recipient)

        # Row A12 has embedded placeholder: "请于 ← 填写 去以下仓库装箱"
        if self.job.loading_date:
            cell_val = ws["A12"].value or ""
            if "← 填写" in str(cell_val):
                ws["A12"].value = str(cell_val).replace("← 填写", self.job.loading_date)

        _set_cell(ws, "C14", self.job.warehouse_address)
        _set_cell(ws, "C16", self.job.warehouse_phone)

        # C19: "← 填写           航次  ← 填写" — contains vessel and voyage
        if self.job.vessel_name or self.job.voyage:
            vessel_part = self.job.vessel_name or ""
            voyage_part = self.job.voyage or ""
            _set_cell(ws, "C19", f"{vessel_part}           航次  {voyage_part}")

        _set_cell(ws, "C20", self.job.customs_decl_no)

        # C21: "← 填写           箱量  ← 填写" — container type and qty
        if self.job.container_type_qty:
            _set_cell(ws, "C21", self.job.container_type_qty)

        # C22: "← 填写           目的港：  ← 填写" — transit port and destination
        transit = self.job.transit_port or ""
        dest = self.job.pod or ""
        _set_cell(ws, "C22", f"{transit}           目的港：  {dest}")

        _set_cell(ws, "C23", self.job.etd)
        _set_cell(ws, "C24", self.job.job_no_ref or self.job.job_no)
        _set_cell(ws, "C25", self.job.container_seal_deadline)

        wb.save(path)
        wb.close()
        return path


# ── Orchestrator ──────────────────────────────────────────

def generate_all(job, tables: list[str] | None = None) -> dict[str, Path]:
    """Generate all requested tables and return {name: path} dict.

    Args:
        job: OpsJob ORM instance
        tables: list of table keys, e.g. ["ens", "ics2", "multi_product", "manifest", "loading_notice"]
                defaults to generating all 4 cutoff tables + loading notice
    """
    if tables is None:
        tables = ["ens", "ics2", "manifest", "loading_notice"]

    result: dict[str, Path] = {}

    filler_map = {
        "ens": ENSFiller,
        "ics2": ICS2Filler,
        "multi_product": MultiProductFiller,
        "manifest": ManifestFiller,
        "loading_notice": LoadingNoticeFiller,
    }

    for table_key in tables:
        cls = filler_map.get(table_key)
        if cls is None:
            continue
        try:
            filler = cls(job)
            out_path = filler.fill()
            result[table_key] = out_path
        except Exception as e:
            # Log but don't fail the entire batch
            result[table_key] = None
            print(f"[excel_engine] Failed to generate {table_key}: {e}")

    return result


def make_zip(job_id: int, generated: dict[str, Path]) -> Path:
    """Package generated files into a ZIP archive."""
    out_dir = OUTPUT_DIR / str(job_id)
    zip_path = out_dir / f"截关_{job_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"

    name_map = {
        "ens": "ENS_VGM申报表",
        "ics2": "ICS2舱单数据表",
        "multi_product": "多品名表",
        "manifest": "DEHAM舱单表",
        "loading_notice": "做箱通知",
    }

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for key, filepath in generated.items():
            if filepath and filepath.exists():
                label = name_map.get(key, key)
                arcname = f"{label}.xlsx"
                zf.write(filepath, arcname)

    return zip_path
